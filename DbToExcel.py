import os
import sqlite3
from openpyxl import Workbook 

#Change directory to where the file should be
os.chdir("/Users/joshy/Desktop")

conn = sqlite3.connect("scouting.db")
cur = conn.cursor()
cur.execute("select * from matchTable;")

#All results from matchTable put into dictionary I think
results = cur.fetchall()

#Creates empty dictionary to use for inserting into excel file later
final = []
dict = {}

for index in results:
    #Contains the whole sent string, parse by commas to seperate out into vars that can then be put into columns
    #Put into dictionary per row by id with all shit in it

    #Parse given index
    dict = {
    "Team Id" : index[1],
    "Match_id" : index[2],
    "TeleOp_cs1" : index[3],
    "TeleOp_cf1" : index[4],
    "TeleOp_cs2" : index[5],
    "TeleOp_cf2" : index[6],
    "TeleOp_cs3" : index[7],
    "TeleOp_cf3" : index[8],
    "TeleOp_cs" : index[9],
    "TeleOp_cf" : index[10],
    "TeleOp_hs1" : index[11],
    "TeleOp_hf1" : index[12],
    "TeleOp_hs2" : index[13],
    "TeleOp_hf2" : index[14],
    "TeleOp_hs3" : index[15],
    "TeleOp_hf3" : index[16],
    "TeleOp_hs" : index[17],
    "TeleOp_hf" : index[18],
    "Rocket_h1" : index[19],
    "Rocket_h2" : index[20],
    "Rocket_h3" : index[21],
    "Rocket_c1" : index[22],
    "Rocket_c2" : index[23],
    "Rocket_c3" : index[24],
    "CargoShip_front_c" : index[25],
    "CargoShip_front_h" : index[26],
    "CargoShip_Side_c" : index[27],
    "CargoShip_Side_h" : index[28],
    "Habitat" : index[29],
    "Assisted" : index[30],
    "Level_Climbed" : index[31],
    "You_Assisted" : index[32],
    "Starting_Level" : index[33],
    "Leave_Habitat" : index[34],
    "Defense_Played" : index[35],
    "Notes" : index[36],
    "Defense_Played_On" : index[37]
    }

    final.append(dict)
    #print(dict["Notes"])

#Put into excel document
wb = Workbook()

# grab the active worksheet
ws = wb.active

i = 2

ws['A1'].value = "Team_id"
ws['B1'].value = "Match_id"
ws['C1'].value = "TeleOp_cs1"
ws['D1'].value = "TeleOp_cf1"
ws['E1'].value = "TeleOp_cs2"
ws['F1'].value = "TeleOp_cf2"
ws['G1'].value = "TeleOp_cs3"
ws['H1'].value = "TeleOp_cf3"
ws['I1'].value = "TeleOp_cs"
ws['J1'].value = "TeleOp_cf"
ws['K1'].value = "TeleOp_hs1"
ws['L1'].value = "TeleOp_hf1"
ws['M1'].value = "TeleOp_hs2"
ws['N1'].value = "TeleOp_hf2"
ws['O1'].value = "TeleOp_hs3"
ws['P1'].value = "TeleOp_hf3"
ws['Q1'].value = "TeleOp_hs"
ws['R1'].value = "TeleOp_hf"
ws['S1'].value = "Rocket_h1"
ws['T1'].value = "Rocket_h2"
ws['U1'].value = "Rocket_h3"
ws['V1'].value = "Rocket_c1"
ws['W1'].value = "Rocket_c2"
ws['X1'].value = "Rocket_c3"
ws['Y1'].value = "CargoShip_front_c"
ws['Z1'].value = "CargoShip_front_h"
ws['AA1'].value = "CargoShip_Side_c"
ws['AB1'].value = "CargoShip_Side_h"
ws['AC1'].value = "Habitat"
ws['AD1'].value = "Assisted"
ws['AE1'].value = "Level_Climbed"
ws['AF1'].value = "You_Assisted"
ws['AG1'].value = "Starting_Level"
ws['AH1'].value = "Leave_Habitat"
ws['AI1'].value = "Defense_Played"
ws['AJ1'].value = "Notes"
ws['AK1'].value = "Defense_Played_On"


for index in final:
    ws['A'+ str(i)].value = index["Team Id"]
    ws['B'+ str(i)].value = index["Match_id"]
    ws['C'+ str(i)].value = index["TeleOp_cs1"]
    ws['D'+ str(i)].value = index["TeleOp_cf1"]
    ws['E'+ str(i)].value = index["TeleOp_cs2"]
    ws['F'+ str(i)].value = index["TeleOp_cf2"]
    ws['G'+ str(i)].value = index["TeleOp_cs3"]
    ws['H'+ str(i)].value = index["TeleOp_cf3"]
    ws['I'+ str(i)].value = index["TeleOp_cs"]
    ws['J'+ str(i)].value = index["TeleOp_cf"]
    ws['K'+ str(i)].value = index["TeleOp_hs1"]
    ws['L'+ str(i)].value = index["TeleOp_hf1"]
    ws['M'+ str(i)].value = index["TeleOp_hs2"]
    ws['N'+ str(i)].value = index["TeleOp_hf2"]
    ws['O'+ str(i)].value = index["TeleOp_hs3"]
    ws['P'+ str(i)].value = index["TeleOp_hf3"]
    ws['Q'+ str(i)].value = index["TeleOp_hs"]
    ws['R'+ str(i)].value = index["TeleOp_hf"]
    ws['S'+ str(i)].value = index["Rocket_h1"]
    ws['T'+ str(i)].value = index["Rocket_h2"]
    ws['U'+ str(i)].value = index["Rocket_h3"]
    ws['V'+ str(i)].value = index["Rocket_c1"]
    ws['W'+ str(i)].value = index["Rocket_c2"]
    ws['X'+ str(i)].value = index["Rocket_c3"]
    ws['Y'+ str(i)].value = index["CargoShip_front_c"]
    ws['Z'+ str(i)].value = index["CargoShip_front_h"]
    ws['AA'+ str(i)].value = index["CargoShip_Side_c"]
    ws['AB'+ str(i)].value = index["CargoShip_Side_h"]
    ws['AC'+ str(i)].value = index["Habitat"]
    ws['AD'+ str(i)].value = index["Assisted"]
    ws['AE'+ str(i)].value = index["Level_Climbed"]
    ws['AF'+ str(i)].value = index["You_Assisted"]
    ws['AG'+ str(i)].value = index["Starting_Level"]
    ws['AH'+ str(i)].value = index["Leave_Habitat"]
    ws['AI'+ str(i)].value = index["Defense_Played"]
    ws['AJ'+ str(i)].value = index["Notes"] 
    ws['AK'+ str(i)].value = index["Defense_Played_On"]

    i += 1

# Save the file
wb.save("scoutingData.xlsx")
wb.close()
#Close connections
cur.close()
conn.close()